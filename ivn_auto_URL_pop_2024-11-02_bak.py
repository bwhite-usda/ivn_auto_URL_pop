import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the Excel workbook and worksheet
file_path = 'ivntest.xlsx'
wb = load_workbook(file_path)
ws = wb.active

# Load data into a pandas DataFrame for easier processing
df = pd.read_excel(file_path)

# Identify columns for Enabling and Dependent components and URLs
enabling_col = 'Enabling Component'
enabling_url_col = 'Enabling Component URL'
dependent_col = 'Dependent Component'
dependent_url_col = 'Dependent Component URL'

# Using dictionaries to store the first non-blank URL for each unique component
enabling_urls = {}
dependent_urls = {}

# Collect the first non-blank URL for each component across all rows
for i, row in df.iterrows():
    # Store Enabling Component URL if present
    enabling_component = row[enabling_col]
    enabling_url = row[enabling_url_col]
    if pd.notnull(enabling_url):
        enabling_urls[enabling_component] = enabling_url
    
    # Store Dependent Component URL if present
    dependent_component = row[dependent_col]
    dependent_url = row[dependent_url_col]
    if pd.notnull(dependent_url):
        dependent_urls[dependent_component] = dependent_url

# Define the highlight fill style
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Fill in missing URLs by looking up the first stored URL for each component
for i, row in df.iterrows():
    enabling_component = row[enabling_col]
    dependent_component = row[dependent_col]
    
    # Fill in Enabling Component URL if missing and highlight the cell
    if pd.isnull(row[enabling_url_col]) and enabling_component in enabling_urls:
        df.at[i, enabling_url_col] = enabling_urls[enabling_component]
        ws.cell(row=i+2, column=df.columns.get_loc(enabling_url_col) + 1, value=enabling_urls[enabling_component]).fill = highlight_fill
    
    # Fill in Dependent Component URL if missing and highlight the cell
    if pd.isnull(row[dependent_url_col]) and dependent_component in dependent_urls:
        df.at[i, dependent_url_col] = dependent_urls[dependent_component]
        ws.cell(row=i+2, column=df.columns.get_loc(dependent_url_col) + 1, value=dependent_urls[dependent_component]).fill = highlight_fill

# Save the workbook with the updates and highlighted cells
wb.save('ivn_auto_URL_populated_highlighted_v2.xlsx')
print("Dataset updated and saved as ivn_auto_URL_populated_highlighted_v2.xlsx with highlighted cells.")
